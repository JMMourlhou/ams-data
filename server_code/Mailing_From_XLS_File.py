import anvil.server
import pandas as pd
import anvil.tables as tables
from anvil.tables import app_tables

@anvil.server.callable
def xls_file_reader(xls_file_name):

  with open(xls_file_name, "r") as f:
    df = pd.read_excel(f)
    for d in df.to_dict(orient="records"):
        # d is now a dict of {columnname -> value} for this row
        # We use Python's **kwargs syntax to pass the whole dict as keyword arguments
        app_tables.from_xls.add_row(**d)







"""
import openpyxl
from openpyxl import load_workbook

@anvil.server.callable
def xls_file_reader(xls_file_name):
    wb = openpyxl.load_workbook(xls_file_name)
    sheet = wb[wb.sheetnames[0]]
    print("feuille 1:", sheet)
    cell=sheet(["M2"])
    return cell
"""